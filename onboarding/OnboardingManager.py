
import openpyxl
from grade.models import Grade


def load_data(path):
    """
        Returns a list of all the data contained in  a given .xlsx files
    """

    wb = openpyxl.load_workbook(path)

    sheets = wb.sheetnames

    # create an empty list to store all the rows
    rows_list = []

    for sheet in sheets:
        current_sheet = wb[sheet]

        # iterate over the rows in the sheet
        for row in range(1, current_sheet.max_row + 1):
            # create an empty list to store the contents of the row
            row_list = []

            # iterate over the columns in the sheet
            for col in range(1, current_sheet.max_column + 1):
                # append the value in the cell to the list
                row_list.append(current_sheet.cell(row, col).value)

            # append the row list to the rows list
            rows_list.append(row_list)

    return rows_list


def parse_row(row):
    """Returns a dictionary containing key/value pairs"""

    full_names = row[0]
    balance_brought_forward = row[1]
    grade = row[2]
    active = row[3]
    year = row[4]
    term = row[5]
    full_names_list = full_names.split(' ')
    first_name = full_names_list[0]
    last_name = full_names_list[-1]
    balance_brought_forward = int(balance_brought_forward)
    grade = Grade.objects.get(title=grade)
    if active == 'TRUE' or active == 'True':
        status = True
        active = bool(status)
    if active == 'FALSE' or active == 'False':
        status = False
        active = bool(status)

    from academic_calendar.models import Year, Term, TermNumbers
    term_number = TermNumbers.objects.get(term=int(term))
    year = Year.objects.get(year=int(year))
    term = Term.objects.get(term=term_number, year=year)

    my_obj = {
        'first_name': first_name,
        'last_name': last_name,
        'balance_brought_forward': balance_brought_forward,
        'grade': grade,
        'active': active,
        'onboarding_year': year,
        'onboarding_term': term
    }
    return my_obj


def register_student(my_obj):
    from student.models import Student
    print("Registering: {}".format(my_obj['first_name']))
    student = Student.objects.create(
        first_name=my_obj['first_name'],
        last_name=my_obj['last_name'],
        grade_admitted_to=my_obj['grade'],
        active=my_obj['active'],
        balance_brought_forward=my_obj['balance_brought_forward'],
        onboarding_year=my_obj['onboarding_year'],
        onboarding_term=my_obj['onboarding_term']
    )
    print("Complete.")
    return student


def register_and_invoice_balance_brought_forward(my_obj):
    """invoices each student with a balance brought forward bill."""

    # get "student_obj
    student = register_student(my_obj)
    print("Registering and invoicing {}".format(student.first_name))
    from invoice.models import Invoice
    from invoice.models import Item as InvoiceItem
    from fees_structure.models import BillingItem

    # Create an invoice for the student
    invoice = Invoice.objects.create(
        student=student,
        year=student.onboarding_year,
        term=student.onboarding_term,
        grade=student.current_grade
    )
    # get the balance brought forward item
    from item.models import Item as SalesItem
    sales_item = SalesItem.objects.get(name='Balance Brought Forward')
    # create a billing item for balance brought forward
    billing_item = BillingItem.objects.create(
        item=sales_item,
        amount=student.balance_brought_forward,
        visible=False
    )

    # create invoice item
    InvoiceItem.objects.create(
        billing_item=billing_item,
        invoice=invoice
    )

    print("Complete...")
    return True


def __main__():
    from onboarding import OnboardingManager
    from invoice.InvoiceManager import InvoiceManager
    inv_man = InvoiceManager()
    data = OnboardingManager.load_data(
        'Kings Educational Centre Data Migration File.xlsx')
    for row in data[1:]:
        OnboardingManager.register_and_invoice_balance_brought_forward(
            OnboardingManager.parse_row(row))
